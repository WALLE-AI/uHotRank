"""Article service for business logic."""

import logging
from typing import List, Optional, Dict, Any
from datetime import datetime

from backend.db.elasticsearch_client import ArticleRepository
from backend.schemas.article import (
    ArticleListResponse,
    ArticleDetail,
    ArticleListItem,
    SearchRequest,
    SearchResponse,
)

logger = logging.getLogger(__name__)


class ArticleService:
    """文章服务层 - 封装文章相关的业务逻辑"""

    def __init__(self, repository: ArticleRepository):
        """
        初始化文章服务

        Args:
            repository: 文章数据仓库实例
        """
        self.repository = repository

    def get_articles(
        self, page: int = 1, size: int = 20, sort_by: Optional[str] = None
    ) -> ArticleListResponse:
        """
        获取文章列表（分页）

        Args:
            page: 页码，从1开始
            size: 每页数量
            sort_by: 排序字段，支持 "publish_date", "scraped_at", "-publish_date", "-scraped_at"

        Returns:
            ArticleListResponse: 分页的文章列表响应
        """
        try:
            # 计算偏移量
            from_ = (page - 1) * size

            # 构建排序规则
            sort = []
            if sort_by:
                # 处理降序标记
                if sort_by.startswith("-"):
                    field = sort_by[1:]
                    order = "desc"
                else:
                    field = sort_by
                    order = "asc"

                # 映射字段名
                field_mapping = {
                    "publish_date": "publish_date",
                    "scraped_at": "scraped_at",
                    "created_at": "scraped_at",  # 兼容前端
                }

                es_field = field_mapping.get(field, "scraped_at")
                sort.append({es_field: {"order": order}})
            else:
                # 默认按爬取时间降序
                sort.append({"scraped_at": {"order": "desc"}})

            # 查询 Elasticsearch
            result = self.repository.search(
                query={"match_all": {}}, size=size, from_=from_, sort=sort
            )

            # 提取文档
            hits = result.get("hits", {}).get("hits", [])
            total = result.get("hits", {}).get("total", {}).get("value", 0)

            # 转换为列表项模型
            articles = []
            for hit in hits:
                source = hit["_source"]
                article = ArticleListItem(
                    id=hit["_id"],
                    url=source.get("original_url", "") or source.get("tophub_url", ""),
                    title=source.get("title", ""),
                    category=source.get("category", ""),
                    published_time=source.get("publish_date", ""),
                    summary=source.get("content_analysis", {}).get("summary"),
                    sentiment=source.get("content_analysis", {}).get("sentiment"),
                )
                articles.append(article)

            return ArticleListResponse(
                articles=articles, total=total, page=page, size=size
            )

        except Exception as e:
            logger.error(f"获取文章列表失败: {e}")
            raise

    def get_article_by_id(self, article_id: str) -> Optional[ArticleDetail]:
        """
        根据 ID 获取文章详情

        Args:
            article_id: 文章 ID

        Returns:
            ArticleDetail: 文章详情，不存在返回 None
        """
        try:
            doc = self.repository.get_document(article_id)

            if not doc:
                return None

            # 转换为详情模型
            article = ArticleDetail(
                id=article_id,
                url=doc.get("original_url", ""),
                title=doc.get("title", ""),
                category=doc.get("category", ""),
                published_time=doc.get("publish_date", ""),
                content=doc.get("content") or None,  # Use None instead of empty string
                tech_detection=doc.get("tech_detection"),
                content_analysis=doc.get("content_analysis"),
                created_at=doc.get("scraped_at", ""),
            )

            return article

        except Exception as e:
            logger.error(f"获取文章详情失败: {e}")
            raise

    def search_articles(self, request: SearchRequest) -> SearchResponse:
        """
        搜索文章

        Args:
            request: 搜索请求参数

        Returns:
            SearchResponse: 搜索结果响应
        """
        try:
            # 计算偏移量
            from_ = (request.page - 1) * request.size

            # 构建查询条件
            must_conditions = []
            filter_conditions = []

            # 关键词搜索
            if request.query:
                must_conditions.append(
                    {
                        "multi_match": {
                            "query": request.query,
                            "fields": ["title^3", "content", "content_analysis.summary^2"],
                            "type": "best_fields",
                        }
                    }
                )

            # 过滤条件
            if request.filters:
                # 分类过滤
                if request.filters.categories:
                    filter_conditions.append(
                        {"terms": {"content_analysis.category.keyword": request.filters.categories}}
                    )

                # 情感过滤
                if request.filters.sentiments:
                    filter_conditions.append(
                        {"terms": {"content_analysis.sentiment.keyword": request.filters.sentiments}}
                    )

                # 日期范围过滤
                if request.filters.date_from or request.filters.date_to:
                    date_range = {}
                    if request.filters.date_from:
                        date_range["gte"] = request.filters.date_from
                    if request.filters.date_to:
                        date_range["lte"] = request.filters.date_to

                    filter_conditions.append({"range": {"publish_date": date_range}})

                # 仅技术文章
                if request.filters.tech_only:
                    filter_conditions.append(
                        {"term": {"tech_detection.is_tech_related": True}}
                    )

            # 构建完整查询
            if must_conditions or filter_conditions:
                query = {"bool": {}}
                if must_conditions:
                    query["bool"]["must"] = must_conditions
                if filter_conditions:
                    query["bool"]["filter"] = filter_conditions
            else:
                query = {"match_all": {}}

            # 构建排序规则
            sort = []
            if request.sort_by:
                if request.sort_by.startswith("-"):
                    field = request.sort_by[1:]
                    order = "desc"
                else:
                    field = request.sort_by
                    order = "asc"

                field_mapping = {
                    "publish_date": "publish_date",
                    "scraped_at": "scraped_at",
                    "relevance": "_score",
                }

                es_field = field_mapping.get(field, "_score")
                if es_field == "_score":
                    sort.append("_score")
                else:
                    sort.append({es_field: {"order": order}})
            elif request.query:
                # 有关键词时默认按相关性排序
                sort.append("_score")
            else:
                # 无关键词时默认按时间降序
                sort.append({"scraped_at": {"order": "desc"}})

            # 执行搜索
            result = self.repository.search(
                query=query, size=request.size, from_=from_, sort=sort
            )

            # 提取文档
            hits = result.get("hits", {}).get("hits", [])
            total = result.get("hits", {}).get("total", {}).get("value", 0)

            # 转换为列表项模型
            articles = []
            for hit in hits:
                source = hit["_source"]
                article = ArticleListItem(
                    id=hit["_id"],
                    url=source.get("original_url", "") or source.get("tophub_url", ""),
                    title=source.get("title", ""),
                    category=source.get("category", ""),
                    published_time=source.get("publish_date", ""),
                    summary=source.get("content_analysis", {}).get("summary"),
                    sentiment=source.get("content_analysis", {}).get("sentiment"),
                )
                articles.append(article)

            return SearchResponse(
                articles=articles, total=total, page=request.page, size=request.size
            )

        except Exception as e:
            logger.error(f"搜索文章失败: {e}")
            raise

    def export_articles(
        self, format: str, fields: List[str], filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        导出文章数据

        Args:
            format: 导出格式 (json, csv, excel)
            fields: 要导出的字段列表
            filters: 过滤条件（可选）

        Returns:
            bytes: 导出的文件内容
        """
        try:
            # 构建查询条件（基于过滤器）
            query = {"match_all": {}}
            if filters:
                filter_conditions = []

                if filters.get("categories"):
                    filter_conditions.append(
                        {"terms": {"content_analysis.category.keyword": filters["categories"]}}
                    )

                if filters.get("sentiments"):
                    filter_conditions.append(
                        {"terms": {"content_analysis.sentiment.keyword": filters["sentiments"]}}
                    )

                if filters.get("date_from") or filters.get("date_to"):
                    date_range = {}
                    if filters.get("date_from"):
                        date_range["gte"] = filters["date_from"]
                    if filters.get("date_to"):
                        date_range["lte"] = filters["date_to"]
                    filter_conditions.append({"range": {"publish_date": date_range}})

                if filters.get("tech_only"):
                    filter_conditions.append(
                        {"term": {"tech_detection.is_tech_related": True}}
                    )

                if filter_conditions:
                    query = {"bool": {"filter": filter_conditions}}

            # 获取所有匹配的文章（限制最大数量）
            max_export_size = 10000
            result = self.repository.search(query=query, size=max_export_size, from_=0)

            hits = result.get("hits", {}).get("hits", [])

            # 提取数据
            articles_data = []
            for hit in hits:
                source = hit["_source"]
                article_data = {"id": hit["_id"]}

                # 根据指定字段提取数据
                field_mapping = {
                    "title": "title",
                    "category": "category",
                    "url": "original_url",
                    "published_time": "publish_date",
                    "content": "content",
                    "summary": "content_analysis.summary",
                    "sentiment": "content_analysis.sentiment",
                    "keywords": "content_analysis.keywords",
                    "topics": "content_analysis.topics",
                    "tech_related": "tech_detection.is_tech_related",
                    "tech_categories": "tech_detection.categories",
                }

                for field in fields:
                    if field in field_mapping:
                        es_field = field_mapping[field]
                        # 处理嵌套字段
                        if "." in es_field:
                            parts = es_field.split(".")
                            value = source
                            for part in parts:
                                value = value.get(part, {}) if isinstance(value, dict) else None
                                if value is None:
                                    break
                            article_data[field] = value
                        else:
                            article_data[field] = source.get(es_field)

                articles_data.append(article_data)

            # 根据格式导出
            if format == "json":
                import json

                return json.dumps(articles_data, ensure_ascii=False, indent=2).encode(
                    "utf-8"
                )

            elif format == "csv":
                import csv
                import io

                output = io.StringIO()
                if articles_data:
                    writer = csv.DictWriter(output, fieldnames=fields)
                    writer.writeheader()
                    for article in articles_data:
                        # 处理列表类型的字段
                        row = {}
                        for field in fields:
                            value = article.get(field)
                            if isinstance(value, list):
                                row[field] = ", ".join(str(v) for v in value)
                            else:
                                row[field] = value
                        writer.writerow(row)

                return output.getvalue().encode("utf-8-sig")  # 使用 BOM 以支持 Excel

            elif format == "excel":
                try:
                    from openpyxl import Workbook
                    from openpyxl.utils import get_column_letter
                    import io

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Articles"

                    # 写入表头
                    for col_idx, field in enumerate(fields, 1):
                        ws.cell(row=1, column=col_idx, value=field)

                    # 写入数据
                    for row_idx, article in enumerate(articles_data, 2):
                        for col_idx, field in enumerate(fields, 1):
                            value = article.get(field)
                            if isinstance(value, list):
                                value = ", ".join(str(v) for v in value)
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    # 调整列宽
                    for col_idx in range(1, len(fields) + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 20

                    # 保存到字节流
                    output = io.BytesIO()
                    wb.save(output)
                    return output.getvalue()

                except ImportError:
                    logger.error("openpyxl 未安装，无法导出 Excel 格式")
                    raise ValueError("Excel 导出功能需要安装 openpyxl 库")

            else:
                raise ValueError(f"不支持的导出格式: {format}")

        except Exception as e:
            logger.error(f"导出文章失败: {e}")
            raise
