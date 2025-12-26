"""Export utility functions for article data."""

import csv
import json
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def select_fields(articles: List[Dict[str, Any]], fields: List[str]) -> List[Dict[str, Any]]:
    """
    Select specific fields from article data.
    
    Args:
        articles: List of article dictionaries
        fields: List of field names to include
        
    Returns:
        List of dictionaries with only selected fields
    """
    if not fields:
        return articles
    
    selected_articles = []
    for article in articles:
        selected_article = {}
        for field in fields:
            # Handle nested fields with dot notation (e.g., "content_analysis.summary")
            if '.' in field:
                parts = field.split('.')
                value = article
                for part in parts:
                    if isinstance(value, dict):
                        value = value.get(part)
                    else:
                        value = None
                        break
                selected_article[field] = value
            else:
                selected_article[field] = article.get(field)
        selected_articles.append(selected_article)
    
    return selected_articles


def flatten_article(article: Dict[str, Any]) -> Dict[str, Any]:
    """
    Flatten nested article structure for CSV/Excel export.
    
    Args:
        article: Article dictionary with nested structures
        
    Returns:
        Flattened dictionary
    """
    flattened = {}
    
    # Basic fields
    for key in ['id', 'url', 'title', 'category', 'published_time', 'created_at']:
        flattened[key] = article.get(key, '')
    
    # Content (truncate for readability)
    content = article.get('content', '')
    flattened['content'] = content[:500] + '...' if len(content) > 500 else content
    
    # Tech detection fields
    tech_detection = article.get('tech_detection', {})
    if tech_detection:
        flattened['is_tech_related'] = tech_detection.get('is_tech_related', False)
        flattened['tech_categories'] = ', '.join(tech_detection.get('categories', []))
        flattened['tech_confidence'] = tech_detection.get('confidence', 0.0)
        flattened['tech_keywords'] = ', '.join(tech_detection.get('matched_keywords', []))
    
    # Content analysis fields
    content_analysis = article.get('content_analysis', {})
    if content_analysis:
        flattened['keywords'] = ', '.join(content_analysis.get('keywords', []))
        flattened['topics'] = ', '.join(content_analysis.get('topics', []))
        flattened['summary'] = content_analysis.get('summary', '')
        flattened['sentiment'] = content_analysis.get('sentiment', 'neutral')
        flattened['analysis_category'] = content_analysis.get('category', '')
        
        # Entities
        entities = content_analysis.get('entities', [])
        if entities:
            entity_names = [e.get('name', '') for e in entities]
            flattened['entities'] = ', '.join(entity_names)
    
    return flattened


def export_to_json(
    articles: List[Dict[str, Any]],
    fields: Optional[List[str]] = None,
    pretty: bool = True
) -> bytes:
    """
    Export articles to JSON format.
    
    Args:
        articles: List of article dictionaries
        fields: Optional list of fields to include (None = all fields)
        pretty: Whether to format JSON with indentation
        
    Returns:
        JSON data as bytes
    """
    # Select fields if specified
    if fields:
        articles = select_fields(articles, fields)
    
    # Convert to JSON
    indent = 2 if pretty else None
    json_str = json.dumps(articles, ensure_ascii=False, indent=indent)
    
    return json_str.encode('utf-8')


def export_to_csv(
    articles: List[Dict[str, Any]],
    fields: Optional[List[str]] = None
) -> bytes:
    """
    Export articles to CSV format.
    
    Args:
        articles: List of article dictionaries
        fields: Optional list of fields to include (None = all fields)
        
    Returns:
        CSV data as bytes
    """
    if not articles:
        return b''
    
    # Flatten articles for CSV export
    flattened_articles = [flatten_article(article) for article in articles]
    
    # Select fields if specified
    if fields:
        flattened_articles = select_fields(flattened_articles, fields)
    
    # Get all unique field names
    if flattened_articles:
        fieldnames = list(flattened_articles[0].keys())
    else:
        fieldnames = []
    
    # Write to CSV
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(flattened_articles)
    
    return output.getvalue().encode('utf-8')


def export_to_excel(
    articles: List[Dict[str, Any]],
    fields: Optional[List[str]] = None
) -> bytes:
    """
    Export articles to Excel format using openpyxl.
    
    Args:
        articles: List of article dictionaries
        fields: Optional list of fields to include (None = all fields)
        
    Returns:
        Excel file data as bytes
    """
    if not articles:
        # Return empty workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Articles"
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # Flatten articles for Excel export
    flattened_articles = [flatten_article(article) for article in articles]
    
    # Select fields if specified
    if fields:
        flattened_articles = select_fields(flattened_articles, fields)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"
    
    # Get field names
    if flattened_articles:
        fieldnames = list(flattened_articles[0].keys())
    else:
        fieldnames = []
    
    # Style for header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Write header row
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=fieldname)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data rows
    for row_idx, article in enumerate(flattened_articles, start=2):
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            value = article.get(fieldname, '')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
